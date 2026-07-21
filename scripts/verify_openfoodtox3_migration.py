#!/usr/bin/env python3
"""Verify the tracked OpenFoodTox 3.0 migration and packaged runtime assets."""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from jsonschema import Draft202012Validator

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))

from dietary_mcp.models import ReferenceValueRecord  # noqa: E402

SOURCE_ID = "efsa.openfoodtox"
LEGACY_SOURCE_ID = "efsa.openfoodtox.2023_snapshot"
SOURCE_DOI = "10.5281/zenodo.19388272"
SOURCE_MD5 = "445fc05a6a421634df822d14131a7d83"
SOURCE_SHA256 = "5181661e921651d4087b544981e1e5a63b99532844eac0112732806a00a85eda"
SCHEMA_SHA256 = "cfa197b4b57a2517eebaac67938601a07137e44942de3ee11194bc477474ac10"

CURATED_PRECEDENCE_KEYS = {
    "acetamiprid",
    "acrylamide",
    "bisphenol_a",
    "cadmium",
    "difenoconazole",
    "ethiprole",
    "glufosinate",
    "glyphosate",
    "imidacloprid",
    "inorganic_arsenic",
    "inorganic_mercury",
    "lead",
    "methylmercury",
    "oxamyl",
    "pfas_4_group",
    "tebuconazole",
    "tetraconazole",
}


class MigrationVerificationError(RuntimeError):
    """Raised when tracked migration evidence violates a release invariant."""


def _load_json(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise MigrationVerificationError(f"expected JSON object: {path}")
    return payload


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise MigrationVerificationError(message)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _workbook_headers(
    repo_root: Path,
    inventory: dict[str, Any],
) -> dict[str, set[str]]:
    sheets = inventory.get("workbook", {}).get("sheets")
    _require(isinstance(sheets, list), "source inventory has no workbook sheets")
    headers_by_sheet: dict[str, set[str]] = {}
    for sheet in sheets:
        name = sheet.get("name")
        headers = sheet.get("headers")
        _require(isinstance(name, str) and name, "source inventory has unnamed sheet")
        _require(isinstance(headers, list), f"source inventory sheet {name} has no headers")
        headers_by_sheet[name] = {
            str(header) for header in headers if isinstance(header, str) and header
        }

    workbook_path = (
        repo_root / "tmp" / "regulatory_sources" / "openfoodtox_3" / "OFT3.0 export repository.xlsx"
    )
    if workbook_path.exists():
        _require(_sha256(workbook_path) == SOURCE_SHA256, "local workbook SHA-256 drift")
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for sheet_name, expected_headers in headers_by_sheet.items():
                _require(sheet_name in workbook.sheetnames, f"local workbook lacks {sheet_name}")
                first_row = next(
                    workbook[sheet_name].iter_rows(
                        min_row=1,
                        max_row=1,
                        values_only=True,
                    )
                )
                actual_headers = {
                    str(header) for header in first_row if header is not None and str(header)
                }
                _require(
                    actual_headers == expected_headers,
                    f"tracked header inventory drift for {sheet_name}",
                )
        finally:
            workbook.close()
    return headers_by_sheet


def _verify_field_path(
    headers_by_sheet: dict[str, set[str]],
    *,
    source_sheet: Any,
    field_path: Any,
    label: str,
    required: bool = False,
) -> None:
    _require(
        isinstance(source_sheet, str) and source_sheet in headers_by_sheet,
        f"{label} references an unknown source sheet",
    )
    if field_path is None:
        _require(not required, f"{label} is missing a required field path")
        return
    _require(
        isinstance(field_path, str) and bool(field_path),
        f"{label} has an invalid field path",
    )
    _require(
        field_path in headers_by_sheet[source_sheet],
        f"{label} field path does not resolve to a workbook header: {field_path}",
    )


def _verify_source_fingerprints(repo_root: Path) -> dict[str, set[str]]:
    review_root = repo_root / "docs" / "reviews"
    inventory = _load_json(review_root / "openfoodtox-3-source-inventory.json")
    extraction = _load_json(review_root / "openfoodtox-3-extraction-summary.json")
    candidate = _load_json(review_root / "openfoodtox-3-candidate-summary.json")
    for label, payload in (
        ("inventory", inventory),
        ("extraction", extraction),
        ("candidate", candidate),
    ):
        source = payload["source"]
        _require(source["doi"] == SOURCE_DOI, f"{label} DOI drift")
        _require(source["md5"] == SOURCE_MD5, f"{label} MD5 drift")
        _require(source["sha256"] == SOURCE_SHA256, f"{label} SHA-256 drift")
        schema = payload.get("workbookSchemaSha256") or payload.get("workbook", {}).get(
            "schemaSha256"
        )
        _require(schema == SCHEMA_SHA256, f"{label} workbook schema drift")
    _require(
        extraction.get("humanHealthRecordCount") == 38808,
        "human-health extraction count drift",
    )
    _require(
        candidate.get("referencePointCandidateCount") == 171,
        "reference-point candidate count drift",
    )
    _require(
        candidate.get("referencePointSourceEncodingCounts")
        == {"structured": 133, "unstructured": 38},
        "reference-point source-encoding count drift",
    )
    return _workbook_headers(repo_root, inventory)


def _verify_source_catalog(repo_root: Path) -> None:
    sources: dict[str, dict[str, Any]] = {}
    for path in sorted((repo_root / "defaults" / "v1").glob("source_catalog*.json")):
        catalog = _load_json(path)
        for item in catalog["sources"]:
            source_id = item["sourceId"]
            _require(source_id not in sources, f"duplicate source id {source_id}")
            sources[source_id] = item
    current = sources[SOURCE_ID]
    legacy = sources[LEGACY_SOURCE_ID]
    _require(current["documentStatus"] == "dataset_current", "3.0 source is not current")
    _require(current["effectiveDate"] == "2026-04-30", "3.0 source date drift")
    _require(legacy["documentStatus"] == "superseded", "2.0 source is not superseded")
    _require(legacy["supersededBy"] == [SOURCE_ID], "2.0 supersession link drift")
    _require(LEGACY_SOURCE_ID in current["supersedes"], "3.0 supersedes link drift")
    glyphosate_2023 = sources["efsa.glyphosate.peer_review.2023"]
    glyphosate_2015 = sources["efsa.glyphosate.peer_review.2015"]
    _require(
        glyphosate_2015["documentStatus"] == "superseded",
        "2015 glyphosate source is not historical",
    )
    _require(
        glyphosate_2015["sourceId"] in glyphosate_2023["supersedes"],
        "glyphosate source supersession link drift",
    )
    imidacloprid_2008 = sources["efsa.imidacloprid.peer_review.2008"]
    joint_2013 = sources["efsa.acetamiprid_imidacloprid.dnt.2013"]
    imidacloprid_2019 = sources["efsa.imidacloprid.mrl_review.2019"]
    _require(
        imidacloprid_2008["documentStatus"] == "final_current"
        and not imidacloprid_2008["supersededBy"],
        "2008 imidacloprid context is incorrectly superseded",
    )
    _require(
        not joint_2013["supersedes"],
        "2013 imidacloprid recommendation incorrectly claims supersession",
    )
    _require(
        imidacloprid_2019["documentStatus"] == "final_current"
        and imidacloprid_2019["effectiveDate"] == "2019-01-31",
        "2019 imidacloprid MRL assessment context drift",
    )


def _verify_runtime_and_provenance(
    repo_root: Path,
    headers_by_sheet: dict[str, set[str]],
) -> tuple[int, int]:
    defaults_root = repo_root / "defaults" / "v1"
    pack = _load_json(defaults_root / "reference_values_openfoodtox.json")
    provenance = _load_json(defaults_root / "openfoodtox_reference_value_provenance.json")
    summary = _load_json(repo_root / "docs" / "reviews" / "openfoodtox-3-candidate-summary.json")
    records = pack.get("records")
    provenance_records = provenance.get("records")
    _require(isinstance(records, list), "runtime pack has no records array")
    _require(isinstance(provenance_records, list), "provenance has no records array")
    _require(
        len(records) == summary["emittedRuntimeRecordCount"] == 2417,
        "runtime candidate count drift",
    )
    _require(summary["sourceCandidateCount"] == 2734, "source candidate count drift")
    _require(summary["heldCandidateCount"] == 317, "held candidate count drift")
    _require(
        summary["excludedOperatorSectionRecordCounts"]
        == {
            "acceptableOperatorExposureLevel": 622,
            "acuteAcceptableOperatorExposureLevel": 194,
        },
        "operator-section exclusion count drift",
    )

    record_ids: set[str] = set()
    runtime_substance_keys: set[str] = set()
    for item in records:
        validated = ReferenceValueRecord.model_validate(item)
        _require(validated.record_id not in record_ids, "duplicate OpenFoodTox record ID")
        record_ids.add(validated.record_id)
        runtime_substance_keys.add(validated.substance_key)
        _require(validated.source_ids == [SOURCE_ID], "runtime record source drift")
        _require(validated.database_source_id == SOURCE_ID, "database source drift")
        _require(validated.document_status.value == "dataset_current", "record is not current")
        _require(validated.submission_use.value == "review_required", "record bypasses review")
        _require(validated.source_output_id is None, "synthetic source output ID detected")
        _require(validated.value is not None and validated.value > 0, "invalid runtime value")
        _require(bool(validated.unit), "runtime value has no unit")
        _require(bool(validated.population), "runtime value has no population")
    _require(
        not runtime_substance_keys & CURATED_PRECEDENCE_KEYS,
        "bulk pack overrides a curated-precedence substance",
    )

    provenance_by_id = {item["recordId"]: item for item in provenance_records}
    _require(len(provenance_by_id) == len(provenance_records), "duplicate provenance ID")
    _require(record_ids == set(provenance_by_id), "runtime/provenance identity mismatch")
    for record_id, item in provenance_by_id.items():
        source = item.get("openfoodtox3") or {}
        _require(source.get("recordKey"), f"{record_id} lacks source record key")
        _require(source.get("toxReferenceDocumentUuid"), f"{record_id} lacks tox UUID")
        _require(source.get("substanceUuid"), f"{record_id} lacks substance UUID")
        _require(source.get("bound") in {"lower", "upper"}, f"{record_id} lacks bound")
        _require(
            source.get("sourceFile") == "OFT3.0 export repository.xlsx",
            f"{record_id} lacks source file",
        )
        _require(
            source.get("sourceSheet") == "FLEX_SUM.ToxRefValues", f"{record_id} source sheet drift"
        )
        _verify_field_path(
            headers_by_sheet,
            source_sheet=source.get("sourceSheet"),
            field_path=source.get("sourceFieldPath"),
            label=f"{record_id} source value",
            required=True,
        )
        _verify_field_path(
            headers_by_sheet,
            source_sheet=source.get("sourceSheet"),
            field_path=source.get("sourceUnitFieldPath"),
            label=f"{record_id} source unit",
            required=True,
        )
        _verify_field_path(
            headers_by_sheet,
            source_sheet=source.get("sourceSheet"),
            field_path=source.get("sourceQualifierFieldPath"),
            label=f"{record_id} source qualifier",
        )
        _require(source.get("rawValue") is not None, f"{record_id} lacks raw value")
        _require(source.get("rawUnit"), f"{record_id} lacks raw unit")
        _require(bool(source.get("dossiers")), f"{record_id} lacks dossier provenance")
    lower_arfd_records = [
        item["openfoodtox3"]
        for item in provenance_records
        if item.get("openfoodtox3", {}).get("section") == "acuteReferenceDose"
        and item.get("openfoodtox3", {}).get("bound") == "lower"
    ]
    _require(len(lower_arfd_records) == 564, "lower-bound ARfD count drift")
    _require(
        all(item.get("sourceQualifierFieldPath") is None for item in lower_arfd_records),
        "lower-bound ARfD provenance claims a nonexistent qualifier column",
    )
    return len(records), len(provenance_records)


def _verify_high_impact_gate(
    repo_root: Path,
    headers_by_sheet: dict[str, set[str]],
) -> int:
    review = _load_json(repo_root / "docs" / "reviews" / "openfoodtox-3-high-impact-review.json")
    schema = _load_json(
        repo_root / "docs" / "reviews" / "openfoodtox-3-high-impact-review.schema.json"
    )
    schema_errors = sorted(
        Draft202012Validator(schema).iter_errors(review),
        key=lambda error: list(error.path),
    )
    _require(
        not schema_errors,
        "high-impact review schema violation: "
        + (schema_errors[0].message if schema_errors else "unknown"),
    )
    _require(
        review["releaseGate"] == "human_toxicologist_review_required",
        "high-impact human review gate is missing",
    )
    _require(review["complete"] is True, "high-impact review is not marked complete")
    _require(review["recordCount"] == len(review["records"]) == 16, "review truncation detected")
    _require(
        review["supportStatusCounts"]
        == dict(sorted(Counter(item["supportStatus"] for item in review["records"]).items())),
        "support status counts drift",
    )
    _require(
        review["temporalStatusCounts"]
        == dict(sorted(Counter(item["temporalStatus"] for item in review["records"]).items())),
        "temporal status counts drift",
    )
    canonical = {key: value for key, value in review.items() if key != "contentSha256"}
    expected_content_sha256 = hashlib.sha256(
        json.dumps(
            canonical,
            sort_keys=True,
            separators=(",", ":"),
            ensure_ascii=False,
        ).encode("utf-8")
    ).hexdigest()
    _require(
        review["contentSha256"] == expected_content_sha256,
        "high-impact review content hash drift",
    )
    by_id = {item["curatedRecordId"]: item for item in review["records"]}
    expected = {
        "efsa.acrylamide.neoplastic.bmdl10": "exact_supported_structured",
        "efsa.acrylamide.neurotoxicity.bmdl10": "exact_supported_structured",
        "efsa.inorganic_arsenic.skin_cancer.bmdl05": "supported_after_unit_normalization",
        "efsa.lead.developmental_neurotoxicity.bmdl01": "exact_supported_unstructured",
        "efsa.lead.nephrotoxicity.bmdl10": "exact_supported_unstructured",
        "efsa.openfoodtox.acetamiprid.adi": "exact_supported_structured",
        "efsa.openfoodtox.acetamiprid.arfd": "supported_after_primary_source_unit_correction",
        "efsa.openfoodtox.glyphosate.arfd": "supported_after_primary_source_unit_correction",
        "efsa.openfoodtox.imidacloprid.arfd": "supported_after_primary_source_unit_correction",
        "efsa.openfoodtox.imidacloprid.arfd.2008": "exact_supported_structured",
    }
    for record_id, status in expected.items():
        _require(
            by_id.get(record_id, {}).get("supportStatus") == status,
            f"{record_id} gate drift",
        )
    _require(
        not {
            "conflicting_values",
            "not_found_after_cross_sheet_search",
            "original_output_overrides_dataset",
        }
        & set(review["supportStatusCounts"]),
        "high-impact review retains unresolved support statuses",
    )

    glyphosate = by_id["efsa.openfoodtox.glyphosate.arfd"]
    _require(
        glyphosate["curatedValue"] == 1.5 and glyphosate["curatedUnit"] == "mg/kg bw",
        "current glyphosate ARfD drift",
    )
    glyphosate_2023 = next(
        item for item in glyphosate["openfoodtox3Candidates"] if item["temporalStatus"] == "current"
    )
    _require(
        glyphosate_2023["sourceFieldPath"].endswith("Arfd.upperValue")
        and glyphosate_2023["rawUnit"] == "mg/kg bw/day",
        "glyphosate raw source encoding was not preserved",
    )
    _require(
        glyphosate_2023["normalizedUnit"] == "mg/kg bw"
        and glyphosate_2023["relationshipToCurated"] == "primary_source_unit_correction"
        and glyphosate_2023["unitCorrection"]["authoritySourceId"]
        == "efsa.glyphosate.peer_review.2023",
        "glyphosate authoritative unit correction drift",
    )

    acetamiprid = by_id["efsa.openfoodtox.acetamiprid.adi"]
    _require(
        acetamiprid["assertionStatus"] == "proposed"
        and acetamiprid["regulatoryFollowUpSourceIds"] == ["eu.reg.2025_158"],
        "acetamiprid authority or follow-up semantics drift",
    )
    acetamiprid_current = next(
        item
        for item in acetamiprid["openfoodtox3Candidates"]
        if item["temporalStatus"] == "current"
    )
    _require(
        [item["sourceId"] for item in acetamiprid_current["introducedBy"]]
        == ["efsa.acetamiprid.statement.2024"],
        "acetamiprid assertion-level provenance drift",
    )
    acetamiprid_arfd = by_id["efsa.openfoodtox.acetamiprid.arfd"]
    acetamiprid_arfd_current = next(
        item
        for item in acetamiprid_arfd["openfoodtox3Candidates"]
        if item["temporalStatus"] == "current"
    )
    _require(
        acetamiprid_arfd["curatedUnit"] == "mg/kg bw"
        and acetamiprid_arfd_current["rawUnit"] == "mg/kg bw/day"
        and acetamiprid_arfd_current["normalizedUnit"] == "mg/kg bw"
        and acetamiprid_arfd_current["relationshipToCurated"] == "primary_source_unit_correction",
        "acetamiprid ARfD unit correction drift",
    )

    imidacloprid = by_id["efsa.openfoodtox.imidacloprid.arfd"]
    _require(
        imidacloprid["curatedValue"] == 0.06
        and imidacloprid["curatedUnit"] == "mg/kg bw"
        and imidacloprid["assertionStatus"] == "recommended"
        and not imidacloprid["supersedesRecordIds"]
        and any(
            item["normalizedValue"] == 0.08
            and item["temporalStatus"] == "current"
            and item["relationshipToCurated"] == "different_current_value"
            for item in imidacloprid["openfoodtox3Candidates"]
        ),
        "imidacloprid parallel-context adjudication drift",
    )
    imidacloprid_recommendation = next(
        item for item in imidacloprid["openfoodtox3Candidates"] if item["normalizedValue"] == 0.06
    )
    _require(
        imidacloprid_recommendation["rawUnit"] == "mg/kg bw/day"
        and imidacloprid_recommendation["normalizedUnit"] == "mg/kg bw"
        and imidacloprid_recommendation["relationshipToCurated"]
        == "primary_source_unit_correction",
        "imidacloprid recommendation unit correction drift",
    )
    imidacloprid_retained = by_id["efsa.openfoodtox.imidacloprid.arfd.2008"]
    _require(
        imidacloprid_retained["curatedValue"] == 0.08
        and imidacloprid_retained["assertionStatus"] == "retained_in_mrl_assessment"
        and imidacloprid_retained["regulatoryFollowUpSourceIds"]
        == ["efsa.imidacloprid.mrl_review.2019"],
        "imidacloprid retained MRL context drift",
    )
    for record in review["records"]:
        for candidate in record["openfoodtox3Candidates"]:
            for field_name, required in (
                ("sourceFieldPath", True),
                ("sourceUnitFieldPath", True),
                ("sourceQualifierFieldPath", False),
                ("descriptorFieldPath", False),
            ):
                _verify_field_path(
                    headers_by_sheet,
                    source_sheet=candidate.get("sourceSheet"),
                    field_path=candidate.get(field_name),
                    label=f"{record['curatedRecordId']} candidate {field_name}",
                    required=required,
                )
    return len(review["records"])


def _verify_manifest_and_packaged_assets(repo_root: Path) -> None:
    manifest_path = repo_root / "defaults" / "manifest.json"
    manifest = _load_json(manifest_path)
    expected = {item["path"]: item["sha256"] for item in manifest["files"]}
    for relative_path in (
        "defaults/v1/reference_values.json",
        "defaults/v1/reference_values_openfoodtox.json",
        "defaults/v1/openfoodtox_reference_value_provenance.json",
        "defaults/v1/source_catalog_pesticide_expansion.json",
        "defaults/v1/substance_synonyms.json",
    ):
        source_path = repo_root / relative_path
        _require(
            expected.get(relative_path) == _sha256(source_path), f"manifest drift: {relative_path}"
        )
        packaged_path = repo_root / "src" / "dietary_mcp" / "data" / Path(relative_path)
        _require(
            source_path.read_bytes() == packaged_path.read_bytes(),
            f"packaged drift: {relative_path}",
        )
    packaged_manifest = repo_root / "src" / "dietary_mcp" / "data" / "defaults" / "manifest.json"
    _require(
        manifest_path.read_bytes() == packaged_manifest.read_bytes(), "packaged manifest drift"
    )


def verify_migration(repo_root: Path) -> dict[str, Any]:
    headers_by_sheet = _verify_source_fingerprints(repo_root)
    _verify_source_catalog(repo_root)
    runtime_count, provenance_count = _verify_runtime_and_provenance(
        repo_root,
        headers_by_sheet,
    )
    high_impact_count = _verify_high_impact_gate(repo_root, headers_by_sheet)
    _verify_manifest_and_packaged_assets(repo_root)
    return {
        "status": "ok",
        "sourceId": SOURCE_ID,
        "sourceDoi": SOURCE_DOI,
        "runtimeRecordCount": runtime_count,
        "provenanceRecordCount": provenance_count,
        "highImpactReviewRecordCount": high_impact_count,
        "releaseGate": "human_toxicologist_review_required",
    }


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--repo-root",
        type=Path,
        default=REPO_ROOT,
    )
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    try:
        result = verify_migration(args.repo_root.resolve())
    except MigrationVerificationError as exc:
        raise SystemExit(str(exc)) from exc
    print(json.dumps(result, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""Verify and inventory the pinned OpenFoodTox 3.0 workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

SOURCE_DOI = "10.5281/zenodo.19388272"
SOURCE_RECORD_ID = "19388272"
SOURCE_VERSION = "v7"
SOURCE_PUBLICATION_DATE = "2026-04-30"
SOURCE_FILE_NAME = "OFT3.0 export repository.xlsx"
SOURCE_MD5 = "445fc05a6a421634df822d14131a7d83"
SOURCE_SIZE_BYTES = 22_595_502

DOCUMENT_UUID = "Document UUID"
PARENT_UUID = "Parent UUID"
DOSSIER_UUID = "DOSSIER UUID"
DOCUMENT_LINK_UUID = "DOCUMENT UUID"
SUBSTANCE_REFERENCE = "ReferenceSubstance.ReferenceSubstance"
OPINION_REFERENCE = (
    "HumanHealthHazardCharacteristics.OtherReferenceValues.ReferenceToEFSAOpinion"
)

REQUIRED_COLUMNS: dict[str, tuple[str, ...]] = {
    "DOSSIER": (
        DOCUMENT_UUID,
        "DossierSubject.Name",
        "LiteratureReference.DateOfEvaluation",
        "LiteratureReference.EFSAOutputTitle",
        "LiteratureReference.LinkToPersistentIdentifier",
    ),
    "DOSSIER_DOCS": (DOSSIER_UUID, DOCUMENT_LINK_UUID),
    "REF_SUB": (
        DOCUMENT_UUID,
        "Inventory.CASNumber",
        "MolecularStructuralInfo.InChIKey",
        "MolecularStructuralInfo.InChl",
        "MolecularStructuralInfo.SmilesNotation",
        "ReferenceSubstanceName",
    ),
    "SUB": (DOCUMENT_UUID, "ChemicalName", SUBSTANCE_REFERENCE),
    "LIT": (
        DOCUMENT_UUID,
        "GeneralInfo.Name",
        "GeneralInfo.ReferenceYear",
        "GeneralInfo.Source",
    ),
    "FLEX_SUM.ToxRefValues": (DOCUMENT_UUID, PARENT_UUID, OPINION_REFERENCE),
}


class SourceIntegrityError(RuntimeError):
    """Raised when the workbook does not match its pinned source identity."""


def _file_digest(path: Path, algorithm: str) -> str:
    digest = hashlib.new(algorithm)
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_sha256(payload: Any) -> str:
    encoded = json.dumps(
        payload,
        ensure_ascii=True,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _headers(worksheet: Worksheet) -> list[str | None]:
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(value) if value is not None else None for value in first_row]


def _duplicate_values(values: Iterable[str]) -> list[str]:
    counts = Counter(values)
    return sorted(value for value, count in counts.items() if count > 1)


def _selected_rows(
    worksheet: Worksheet,
    columns: tuple[str, ...],
) -> Iterable[dict[str, Any]]:
    headers = _headers(worksheet)
    positions = {column: headers.index(column) for column in columns}
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        yield {
            column: values[position] if position < len(values) else None
            for column, position in positions.items()
        }


def _strings(values: Iterable[Any]) -> list[str]:
    return [str(value).strip() for value in values if value is not None and str(value).strip()]


def _required_schema_issues(workbook: Any) -> list[str]:
    issues: list[str] = []
    for sheet_name, required_columns in REQUIRED_COLUMNS.items():
        if sheet_name not in workbook.sheetnames:
            issues.append(f"missing required sheet: {sheet_name}")
            continue
        available = set(_headers(workbook[sheet_name]))
        for column in required_columns:
            if column not in available:
                issues.append(f"{sheet_name}: missing required column: {column}")
    return issues


def _join_inventory(workbook: Any) -> dict[str, Any]:
    ref_rows = list(_selected_rows(workbook["REF_SUB"], (DOCUMENT_UUID,)))
    sub_rows = list(_selected_rows(workbook["SUB"], (DOCUMENT_UUID, SUBSTANCE_REFERENCE)))
    dossier_rows = list(_selected_rows(workbook["DOSSIER"], (DOCUMENT_UUID,)))
    dossier_doc_rows = list(
        _selected_rows(workbook["DOSSIER_DOCS"], (DOSSIER_UUID, DOCUMENT_LINK_UUID))
    )
    literature_rows = list(_selected_rows(workbook["LIT"], (DOCUMENT_UUID,)))
    tox_rows = list(
        _selected_rows(
            workbook["FLEX_SUM.ToxRefValues"],
            (DOCUMENT_UUID, PARENT_UUID, OPINION_REFERENCE),
        )
    )

    reference_uuids = set(_strings(row[DOCUMENT_UUID] for row in ref_rows))
    substance_uuids = set(_strings(row[DOCUMENT_UUID] for row in sub_rows))
    dossier_uuids = set(_strings(row[DOCUMENT_UUID] for row in dossier_rows))
    literature_uuids = set(_strings(row[DOCUMENT_UUID] for row in literature_rows))
    tox_document_uuid_rows = _strings(row[DOCUMENT_UUID] for row in tox_rows)
    tox_document_uuids = set(tox_document_uuid_rows)

    substance_reference_uuids = set(
        _strings(row[SUBSTANCE_REFERENCE] for row in sub_rows)
    )
    tox_parent_uuids = set(_strings(row[PARENT_UUID] for row in tox_rows))
    linked_dossier_uuids = set(_strings(row[DOSSIER_UUID] for row in dossier_doc_rows))
    linked_document_uuids = set(
        _strings(row[DOCUMENT_LINK_UUID] for row in dossier_doc_rows)
    )
    opinion_uuids = set(_strings(row[OPINION_REFERENCE] for row in tox_rows))

    dossier_links_by_document: dict[str, set[str]] = defaultdict(set)
    for row in dossier_doc_rows:
        document_uuid = _strings((row[DOCUMENT_LINK_UUID],))
        dossier_uuid = _strings((row[DOSSIER_UUID],))
        if document_uuid and dossier_uuid:
            dossier_links_by_document[document_uuid[0]].add(dossier_uuid[0])

    unresolved = {
        "substanceToReferenceSubstance": sorted(substance_reference_uuids - reference_uuids),
        "toxReferenceValueToSubstance": sorted(tox_parent_uuids - substance_uuids),
        "dossierDocumentToDossier": sorted(linked_dossier_uuids - dossier_uuids),
        "toxReferenceValueToLiterature": sorted(opinion_uuids - literature_uuids),
    }
    unlinked_tox_documents = sorted(tox_document_uuids - linked_document_uuids)
    duplicate_tox_documents = _duplicate_values(tox_document_uuid_rows)
    multi_dossier_tox_documents = sorted(
        document_uuid
        for document_uuid in tox_document_uuids
        if len(dossier_links_by_document[document_uuid]) > 1
    )

    errors = [
        f"{relation}: {len(uuids)} unresolved UUIDs"
        for relation, uuids in unresolved.items()
        if uuids
    ]
    review_items: list[str] = []
    if unlinked_tox_documents:
        review_items.append(
            f"{len(unlinked_tox_documents)} toxicological-reference documents have no dossier link"
        )
    if duplicate_tox_documents:
        review_items.append(
            f"{len(duplicate_tox_documents)} toxicological-reference document UUIDs occur more than once"
        )
    if multi_dossier_tox_documents:
        review_items.append(
            f"{len(multi_dossier_tox_documents)} toxicological-reference documents link to multiple dossiers"
        )

    return {
        "entityCounts": {
            "dossierDocumentUuids": len(dossier_uuids),
            "literatureDocumentUuids": len(literature_uuids),
            "referenceSubstanceDocumentUuids": len(reference_uuids),
            "substanceDocumentUuids": len(substance_uuids),
            "toxReferenceValueRows": len(tox_rows),
            "toxReferenceValueDocumentUuids": len(tox_document_uuids),
        },
        "unresolvedJoins": unresolved,
        "unlinkedToxReferenceValueDocumentUuids": unlinked_tox_documents,
        "duplicateToxReferenceValueDocumentUuids": duplicate_tox_documents,
        "multiDossierToxReferenceValueDocumentCount": len(multi_dossier_tox_documents),
        "errors": errors,
        "reviewItems": review_items,
        "status": "error" if errors else "review_required" if review_items else "ok",
    }


def build_inventory(path: Path, *, expected_md5: str = SOURCE_MD5) -> dict[str, Any]:
    path = path.resolve()
    if not path.is_file():
        raise SourceIntegrityError(f"OpenFoodTox source file does not exist: {path}")

    actual_md5 = _file_digest(path, "md5")  # noqa: S324 - required source checksum
    if actual_md5 != expected_md5:
        raise SourceIntegrityError(
            f"OpenFoodTox source checksum mismatch: expected {expected_md5}, got {actual_md5}"
        )

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        schema_issues = _required_schema_issues(workbook)
        if schema_issues:
            raise SourceIntegrityError("; ".join(schema_issues))

        sheets: list[dict[str, Any]] = []
        schema_payload: list[dict[str, Any]] = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            headers = _headers(worksheet)
            named_headers = [header for header in headers if header is not None]
            sheet = {
                "name": sheet_name,
                "dataRows": max(worksheet.max_row - 1, 0),
                "columns": worksheet.max_column,
                "blankHeaderColumns": len(headers) - len(named_headers),
                "duplicateHeaders": _duplicate_values(named_headers),
                "headers": headers,
            }
            sheets.append(sheet)
            schema_payload.append({"name": sheet_name, "headers": headers})

        join_inventory = _join_inventory(workbook)
    finally:
        workbook.close()

    source_size = path.stat().st_size
    source_review_items: list[str] = []
    if expected_md5 == SOURCE_MD5 and source_size != SOURCE_SIZE_BYTES:
        source_review_items.append(
            f"source size differs from Zenodo metadata: expected {SOURCE_SIZE_BYTES}, got {source_size}"
        )

    return {
        "manifestVersion": "1.0",
        "source": {
            "title": "OpenFoodTox 3.0: EFSA's chemical hazards database",
            "doi": SOURCE_DOI,
            "zenodoRecordId": SOURCE_RECORD_ID,
            "version": SOURCE_VERSION,
            "publicationDate": SOURCE_PUBLICATION_DATE,
            "fileName": path.name,
            "sizeBytes": source_size,
            "md5": actual_md5,
            "sha256": _file_digest(path, "sha256"),
            "license": "CC-BY-4.0",
        },
        "workbook": {
            "sheetCount": len(sheets),
            "schemaSha256": _canonical_sha256(schema_payload),
            "sheets": sheets,
        },
        "joinIntegrity": join_inventory,
        "sourceReviewItems": source_review_items,
        "status": (
            "error"
            if join_inventory["errors"]
            else "review_required"
            if join_inventory["reviewItems"] or source_review_items
            else "ok"
        ),
    }


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Path to the pinned OpenFoodTox 3.0 XLSX")
    parser.add_argument("--output", type=Path, help="Write the deterministic JSON inventory")
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    try:
        inventory = build_inventory(args.workbook)
    except SourceIntegrityError as exc:
        raise SystemExit(str(exc)) from exc

    rendered = json.dumps(inventory, indent=2, ensure_ascii=True) + "\n"
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(rendered, encoding="utf-8")
    else:
        print(rendered, end="")
    return 1 if inventory["status"] == "error" else 0


if __name__ == "__main__":
    raise SystemExit(main())
